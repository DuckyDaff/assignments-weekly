"""
merge_annual.py
---------------
Smart-merge statuses from the department Excel INTO the existing production
annual plan, WITHOUT destroying production-only data.

What it preserves (never touched):
  - section structure & people lists (incl. the מחלקתי section)
  - nominalHours, holidays
  - notes, vehicles, and any other per-day fields
  - days/people that have no value in the Excel

What it does:
  - For every production person, resolves their name in the Excel (alias map
    handles spelling differences; matching is by NAME across all columns, so
    people who moved between sections still get their data).
  - For each day in the requested month range where the Excel has a non-empty
    code, it writes that code onto the person — into slot 2 (statuses2) for
    משמרת-section people when the code is a shift/on-call code, otherwise slot 1
    (statuses). The person is first removed from BOTH slots to avoid duplicates.
  - Empty Excel cells are left as-is (no clearing) unless --mirror is passed.

Usage:
  python scripts/merge_annual.py --xlsx "PATH.xlsx"                 # dry run + report
  python scripts/merge_annual.py --xlsx "PATH.xlsx" --upload        # apply to production
  python scripts/merge_annual.py --xlsx "PATH.xlsx" --month-max 6   # Jan–June only
  python scripts/merge_annual.py --xlsx "PATH.xlsx" --mirror        # also clear empty cells
"""

import sys, json, datetime, argparse
import urllib.request, urllib.error

try:
    import openpyxl
except ImportError:
    sys.exit("Run: pip install openpyxl")

# ── Excel column layout (0-based), same template as import_annual.py ──
SECTION_COLS = {
    "מדור פיקוד ובקרה": [3, 5, 7, 9, 11, 13, 15, 17, 19, 21, 23, 25, 27, 29, 31],
    "מדור תאורת מסלולים": [33, 35, 37, 39],
    "משמרת מסלולים": [48, 50, 52, 54, 56, 58, 60, 62, 64, 66, 68, 70, 72, 74, 76, 78],
}

MONTH_NAMES = {
    "ינואר": 1, "פברואר": 2, "מרץ": 3, "אפריל": 4,
    "מאי": 5, "יוני": 6, "יולי": 7, "אוגוסט": 8,
    "ספטמבר": 9, "אוקטובר": 10, "נובמבר": 11, "דצמבר": 12,
}

SHIFT_FAMILY = {"י", "ל", "Y", "L", "כ", "כש", "כמ", "כמש"}

# Production-name -> Excel-name aliases (spelling differences confirmed by user)
ALIAS_PROD_TO_XLSX = {
    "אלון ואיברנדי": "אלון איברנדי",
    "עומרי סלמן": "עמרי סלמן",
    "איציק אפרמשווילי": "איציק אפרמשוילי",
}

API_URL = "https://assignments-weekly.vercel.app/api/annual"


def cv(ws, r, c):
    """0-based col, 1-based row."""
    v = ws.cell(row=r, column=c + 1).value
    if v is None:
        return ""
    if isinstance(v, str):
        return v.replace("\xa0", " ").strip()
    if isinstance(v, (int, float)):
        return str(int(v)) if v == int(v) else str(v)
    return str(v)


def build_excel_map(xlsx_path, year, month_min, month_max):
    """Return {excel_person_name: {iso: (primary, secondary)}} for the range.

    Each person occupies TWO adjacent Excel columns:
      - primary  (col c)   -> main status (absence/training/work)  -> app slot 1
      - secondary(col c+1) -> shift/on-call code (י/ל/כ/כש/כמ/כמש) -> app slot 2
    Both are captured independently so a day can carry e.g. ח + כ together.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    all_cols = [c for cols in SECTION_COLS.values() for c in cols]

    emap = {}
    for sheet_name in wb.sheetnames[:12]:
        ws = wb[sheet_name]
        title = cv(ws, 1, 0).split("\n")[0].strip()
        mn = MONTH_NAMES.get(title)
        if not mn or mn < month_min or mn > month_max:
            continue

        # person names live in row 3
        col_person = {}
        for c in all_cols:
            name = cv(ws, 3, c)
            if name:
                col_person[c] = name

        for day in range(1, 32):
            row = day + 4
            if cv(ws, row, 0) != str(day):
                continue
            try:
                iso = datetime.date(year, mn, day).isoformat()
            except ValueError:
                continue
            for c, person in col_person.items():
                primary   = cv(ws, row, c)
                secondary = cv(ws, row, c + 1)
                if primary or secondary:
                    emap.setdefault(person, {})[iso] = (primary, secondary)
    return emap


def fetch_prod():
    with urllib.request.urlopen(API_URL, timeout=30) as r:
        return json.loads(r.read().decode())


def resolve_excel_name(prod_name, emap):
    if prod_name in ALIAS_PROD_TO_XLSX:
        return ALIAS_PROD_TO_XLSX[prod_name]
    if prod_name in emap:
        return prod_name
    return None


def date_range(year, month_min, month_max):
    d = datetime.date(year, month_min, 1)
    # last day of month_max
    if month_max == 12:
        end = datetime.date(year, 12, 31)
    else:
        end = datetime.date(year, month_max + 1, 1) - datetime.timedelta(days=1)
    while d <= end:
        yield d.isoformat()
        d += datetime.timedelta(days=1)


def merge(plan, emap, year, month_min, month_max, mirror=False):
    days = plan.setdefault("days", {})
    report = {"set": 0, "cleared": 0, "people": [], "unmapped_prod": [], "by_month": {}}

    range_isos = list(date_range(year, month_min, month_max))

    for sec in plan.get("sections", []):
        for prod_name in sec.get("people", []):
            if "נוסף" in prod_name or "תקן" in prod_name:
                continue  # placeholder column, skip
            xname = resolve_excel_name(prod_name, emap)
            if xname is None:
                report["unmapped_prod"].append(f"{sec['name']} / {prod_name}")
                continue

            person_days = emap.get(xname, {})
            n_set = n_clear = 0
            for iso in range_isos:
                primary, secondary = person_days.get(iso, ("", ""))
                dd = days.get(iso)
                if primary or secondary:
                    if dd is None:
                        dd = days[iso] = {}
                    st1 = dd.setdefault("statuses", {})
                    st2 = dd.setdefault("statuses2", {})
                    # primary col -> slot1, secondary col -> slot2.
                    # If a column is empty, remove that slot so the day mirrors Excel exactly.
                    if primary:
                        st1[prod_name] = primary
                    else:
                        st1.pop(prod_name, None)
                    if secondary:
                        st2[prod_name] = secondary
                    else:
                        st2.pop(prod_name, None)
                    n_set += 1
                    mm = iso[:7]
                    report["by_month"][mm] = report["by_month"].get(mm, 0) + 1
                elif mirror and dd is not None:
                    removed = False
                    if prod_name in dd.get("statuses", {}):
                        dd["statuses"].pop(prod_name, None); removed = True
                    if prod_name in dd.get("statuses2", {}):
                        dd["statuses2"].pop(prod_name, None); removed = True
                    if removed:
                        n_clear += 1

            report["set"] += n_set
            report["cleared"] += n_clear
            report["people"].append((sec["name"], prod_name, xname, n_set, n_clear))

    return plan, report


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--year", type=int, default=2026)
    ap.add_argument("--month-min", type=int, default=1)
    ap.add_argument("--month-max", type=int, default=12)
    ap.add_argument("--mirror", action="store_true",
                    help="also clear cells that are empty in the Excel (faithful mirror)")
    ap.add_argument("--upload", action="store_true")
    ap.add_argument("--report", default="_merge_report.txt")
    args = ap.parse_args()

    print(f"Reading Excel {args.xlsx} (months {args.month_min}-{args.month_max}) ...")
    emap = build_excel_map(args.xlsx, args.year, args.month_min, args.month_max)
    print(f"  Excel people with data: {len(emap)}")

    print("Fetching production annual plan ...")
    plan = fetch_prod()
    print(f"  prod sections: {len(plan.get('sections', []))} | days: {len(plan.get('days', {}))}"
          f" | nominalHours={'nominalHours' in plan} holidays={'holidays' in plan}")

    plan, rep = merge(plan, emap, args.year, args.month_min, args.month_max, args.mirror)

    # Write a detailed UTF-8 report
    import io
    o = io.open(args.report, "w", encoding="utf-8")
    o.write(f"MERGE REPORT  months {args.month_min}-{args.month_max}  mirror={args.mirror}\n")
    o.write(f"cells SET: {rep['set']}   cells CLEARED: {rep['cleared']}\n\n")
    o.write("By month (cells set):\n")
    for mm in sorted(rep["by_month"]):
        o.write(f"  {mm}: {rep['by_month'][mm]}\n")
    o.write("\nPer person (section | prod-name -> excel-name | set | cleared):\n")
    for sname, pname, xname, ns, nc in rep["people"]:
        flag = "" if pname == xname else "  (alias)"
        o.write(f"  {sname} | {pname} -> {xname} | set={ns} clear={nc}{flag}\n")
    if rep["unmapped_prod"]:
        o.write("\nProduction people with NO Excel match (left untouched):\n")
        for x in rep["unmapped_prod"]:
            o.write(f"  {x}\n")
    # excel people not present in any prod section
    prod_names = set()
    for sec in plan.get("sections", []):
        for p in sec.get("people", []):
            prod_names.add(p)
            if p in ALIAS_PROD_TO_XLSX:
                prod_names.add(ALIAS_PROD_TO_XLSX[p])
    dropped = [n for n in emap if n not in prod_names]
    if dropped:
        o.write("\nExcel people NOT in production (their data is DROPPED):\n")
        for n in dropped:
            o.write(f"  {n} ({len(emap[n])} days)\n")
    o.close()
    print(f"Report written to {args.report}")
    print(f"  SET {rep['set']} cells | CLEARED {rep['cleared']} cells")

    if args.upload:
        body = json.dumps(plan, ensure_ascii=False).encode("utf-8")
        print(f"Uploading merged plan ({len(body):,} bytes) to {API_URL} ...")
        req = urllib.request.Request(API_URL, data=body, method="PUT",
                                     headers={"Content-Type": "application/json; charset=utf-8"})
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                print("OK", resp.read().decode())
        except urllib.error.HTTPError as e:
            print(f"HTTP {e.code}: {e.read().decode()}")
            sys.exit(1)
    else:
        print("(dry run — review the report, then add --upload to apply)")


if __name__ == "__main__":
    main()
