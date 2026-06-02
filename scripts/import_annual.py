"""
import_annual.py
----------------
Reads תוכנית עבודה שנתית 2026.xlsx and uploads the annual plan
to Redis via the /api/annual endpoint.

Usage:
  python scripts/import_annual.py              # dry run (prints JSON, no upload)
  python scripts/import_annual.py --upload     # upload to https://assignments-weekly.vercel.app
  python scripts/import_annual.py --url http://localhost:5173 --upload  # local dev
"""

import sys, json, datetime, argparse
import urllib.request, urllib.error

try:
    import openpyxl
except ImportError:
    sys.exit("Run: pip install openpyxl")

# ── Column layout (0-based) ──────────────────────────────────────
SECTION_PEOPLE = [
    {
        "name": "מדור פיקוד ובקרה",
        "cols": [3, 5, 7, 9, 11, 13, 15, 17, 19, 21, 23, 25, 27, 29, 31],
    },
    {
        "name": "מדור תאורת מסלולים",
        "cols": [33, 35, 37, 39],
    },
    {
        "name": "משמרת מסלולים",
        "cols": [48, 50, 52, 54, 56, 58, 60, 62, 64, 66, 68, 70, 72, 74, 76, 78],
    },
]

NOTES_COL      = 41
VEHICLE_DMAX   = 43
VEHICLE_STARIA = 44
COUNT_DAY_COL  = 83
COUNT_NIGHT_COL= 84
WORK_DAY_COL   = 85
WORK_NIGHT_COL = 86

MONTH_NAMES = {
    "ינואר": 1, "פברואר": 2, "מרץ": 3, "אפריל": 4,
    "מאי": 5, "יוני": 6, "יולי": 7, "אוגוסט": 8,
    "ספטמבר": 9, "אוקטובר": 10, "נובמבר": 11, "דצמבר": 12,
}

def cell_val(ws, row, col):
    """0-based col, 1-based row (openpyxl is 1-based internally)."""
    v = ws.cell(row=row, column=col + 1).value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.replace("\xa0", " ").strip()
        return v if v else None
    if isinstance(v, (int, float)):
        return v
    return None

def get_status(ws, row, col):
    """Try the person's own column, then col+1 (merged-cell quirk)."""
    v = cell_val(ws, row, col)
    if v is not None:
        return str(v)
    v = cell_val(ws, row, col + 1)
    return str(v) if v is not None else ""

def str_status(v):
    if v is None or v == "":
        return ""
    # Numeric values (2 = extra day, etc.)
    if isinstance(v, (int, float)):
        return str(int(v)) if v == int(v) else str(v)
    return str(v).strip()

def parse_sheet(ws, year):
    """Return list of day-dicts for this month sheet."""
    # Detect month from sheet title (row 1, col 0)
    title = cell_val(ws, 1, 0) or ""
    month_name = title.split("\n")[0].strip()
    month_num = MONTH_NAMES.get(month_name)
    if month_num is None:
        return None, []

    # Read person names from row 3
    col_to_person = {}
    for sec in SECTION_PEOPLE:
        for c in sec["cols"]:
            name = cell_val(ws, 3, c)
            if name and isinstance(name, str) and name.strip():
                col_to_person[c] = name.strip().replace("\xa0", " ")

    days = []
    for day_num in range(1, 32):
        row = day_num + 4   # day 1 → row 5

        # Check this row actually has a day number matching
        actual_day = cell_val(ws, row, 0)
        if actual_day != day_num:
            continue  # month ended

        try:
            date = datetime.date(year, month_num, day_num)
        except ValueError:
            continue

        statuses = {}
        for c, person in col_to_person.items():
            s = get_status(ws, row, c)
            if s:
                statuses[person] = s

        notes    = str_status(cell_val(ws, row, NOTES_COL)) or ""
        veh_dmax = str_status(cell_val(ws, row, VEHICLE_DMAX)) or ""
        veh_star = str_status(cell_val(ws, row, VEHICLE_STARIA)) or ""
        cnt_day  = cell_val(ws, row, COUNT_DAY_COL)
        cnt_ngt  = cell_val(ws, row, COUNT_NIGHT_COL)
        wrk_day  = str_status(cell_val(ws, row, WORK_DAY_COL)) or ""
        wrk_ngt  = str_status(cell_val(ws, row, WORK_NIGHT_COL)) or ""

        days.append({
            "date": date.isoformat(),
            "statuses": statuses,
            "notes": notes,
            "vehicles": {"DMAX": veh_dmax, "Staria": veh_star},
            "countDay": int(cnt_day) if isinstance(cnt_day, (int, float)) else None,
            "countNight": int(cnt_ngt) if isinstance(cnt_ngt, (int, float)) else None,
            "workDay": wrk_day,
            "workNight": wrk_ngt,
        })

    return month_num, days


def build_annual(xlsx_path, year=2026, month_min=1, month_max=12):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Build sections list (people per section) from first available month sheet
    sections = []
    for sec in SECTION_PEOPLE:
        for sheet_name in wb.sheetnames[:12]:
            ws = wb[sheet_name]
            title = (cell_val(ws, 1, 0) or "").split("\n")[0].strip()
            if title not in MONTH_NAMES:
                continue
            people = []
            for c in sec["cols"]:
                name = cell_val(ws, 3, c)
                if name and isinstance(name, str):
                    name = name.strip().replace("\xa0", " ")
                    if name:
                        people.append(name)
            sections.append({"name": sec["name"], "people": people})
            break

    # Build days dict
    all_days = {}
    for sheet_name in wb.sheetnames[:12]:
        ws = wb[sheet_name]
        title = (cell_val(ws, 1, 0) or "").split("\n")[0].strip()
        if title not in MONTH_NAMES:
            continue
        month_num, days = parse_sheet(ws, year)
        if month_num is None:
            continue
        # Only import months within the requested range (e.g. Jan–June)
        if month_num < month_min or month_num > month_max:
            continue
        for d in days:
            all_days[d["date"]] = {k: v for k, v in d.items() if k != "date"}

    return {
        "year": year,
        "sections": sections,
        "days": all_days,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default="תוכנית עבודה שנתית 2026.xlsx")
    parser.add_argument("--year", type=int, default=2026)
    parser.add_argument("--url", default="https://assignments-weekly.vercel.app")
    parser.add_argument("--upload", action="store_true")
    parser.add_argument("--month-min", type=int, default=1, help="first month to import (1-12)")
    parser.add_argument("--month-max", type=int, default=12, help="last month to import (1-12)")
    args = parser.parse_args()

    print(f"Reading {args.xlsx} ...")
    print(f"Month range: {args.month_min}–{args.month_max}")
    data = build_annual(args.xlsx, args.year, args.month_min, args.month_max)

    day_count = len(data["days"])
    sec_info = ", ".join(f"{s['name']} ({len(s['people'])} אנשים)" for s in data["sections"])
    json_bytes = json.dumps(data, ensure_ascii=False).encode("utf-8")

    print(f"OK {day_count} days | {sec_info}")
    print(f"OK JSON size: {len(json_bytes):,} bytes")

    if args.upload:
        url = f"{args.url.rstrip('/')}/api/annual"
        print(f"Uploading to {url} ...")
        req = urllib.request.Request(
            url,
            data=json_bytes,
            method="PUT",
            headers={"Content-Type": "application/json; charset=utf-8"},
        )
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                body = resp.read().decode()
                print(f"OK Server response: {body}")
        except urllib.error.HTTPError as e:
            print(f"✗ HTTP {e.code}: {e.read().decode()}")
            sys.exit(1)
    else:
        # Dry run — print summary of first 3 days
        for date_key in sorted(data["days"])[:3]:
            d = data["days"][date_key]
            print(f"  {date_key}: {d['statuses']} | notes={d['notes']!r}")
        print("(dry run — add --upload to send to server)")


if __name__ == "__main__":
    main()
